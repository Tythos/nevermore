"""
"""

import os
import datetime
import openpyxl

class Meta(object):
    """Defines a metadata object implicitly attached to all entries upon
       interaction with the database.
    """
    
    def __init__(self):
        """
        """
        self._id = 0
        self._cdts = datetime.datetime.utcnow()
        self._mdts = datetime.datetime.utcnow()
        
    @classmethod
    def isAttached(Cls, entry):
        """Returns true if *entry* instance has all attributes of a Meta object
        """
        attrs = Cls().__dict__.keys()
        for att in attrs:
            if not hasattr(entry, att):
                return False
        return True
        
    def attach(self, entry):
        """
        """
        for k, v in self.__dict__.items():
            setattr(entry, k, v)
    
class DataStore(object):
    """
    """
    
    def __init__(self, file):
        """If the given .XLSX file does not exist, a new workbook is created.
           When the context closes and the object is deleted, it will be saved
           from the stored file path. Changes must be explicitly committed, but
           this can be automated by using a *with* clause.
        """
        self.path = file
        if os.path.isfile(file):
            self.wb = openpyxl.load_workbook(file)
        else:
            self.wb = openpyxl.Workbook()
            
    def __enter__(self):
        """
        """
        return self
            
    def __exit__(self, exc_type, exc_value, traceback):
        """
        """
        self.commit()
        
    def commit(self):
        """Saves the workbook to the file provided to the constructor
        """
        self.wb.save(filename=self.path)
        
    def _getTableName(self, Cls):
        """Pluralizes a class name to determine the corresponding worksheet
        """
        if Cls.__name__[-1] == 's':
            return Cls.__name__ + "es"
        return Cls.__name__ + "s"
        
    def _hasTable(self, table):
        """Returns *True* if the workbook has a worksheet with the given
           pluralized name of a class.
        """
        return table in self.wb.get_sheet_names()
       
    def _addTable(self, Cls):
        """Creates a new "Table" (worksheet) based on the metadata fields and
           the keys of the given class.
        """
        fields = list(Meta().__dict__.keys()) + list(Cls().__dict__.keys())
        name = self._getTableName(Cls)
        ws = self.wb.create_sheet(name)
        for ndx, field in enumerate(fields):
            ws.cell(column=ndx+1, row=1, value=field)
        return ws
        
    def _getNextId(self, ws):
        """Gets the next _id field value in the corresponding table/worksheet
        """
        headers = [col[0].value for col in ws.iter_cols()]
        if '_id' not in headers:
            raise Exception("Unable to locate ID column in worksheet")
        id_ndx = headers.index("_id")
        ids = [row[id_ndx].value for ndx, row in enumerate(ws.iter_rows()) if ndx > 0]
        ndx = 0
        while ndx in ids:
            ndx += 1
        return ndx, len(ids) + 1
        
    def _write(self, ws, entry, row_ndx):
        """
        """
        headers = [col[0].value for col in ws.iter_cols()]
        for k, v in entry.__dict__.items():
            ws.cell(column=headers.index(k)+1, row=row_ndx+1, value=v)
            
    def _get(self, ws):
        """
        """
        header = []
        entries = []
        for i, row in enumerate(ws.rows):
            if i is 0:
                header = [c.value for c in row]
            else:
                entry = {}
                for j, cell in enumerate(row):
                    entry[header[j]] = cell.value
                entries.append(entry)
        return entries
        
    def _mask(self, entries, mask):
        """
        """
        return [entry for ndx, entry in enumerate(entries) if mask[ndx]]
        
    def _filter(self, entries, field, constraint):
        """The *constraint* parameter is either a value with an implicit
           equality constraint or a two-item tuple with 1) an inequality string
           and b) a value for comparison by way of the inequality. Returns the
           filtered subset of entries (a list of dictionaries) matching the
           given constraint.
        """
        mask = [True] * len(entries)
        if type(constraint) is tuple:
            for ndx, entry in enumerate(entries):
                ineq = constraint[0]
                if ineq == '<':
                    mask[ndx] = entry[field] < constraint[1]
                elif ineq == '<=':
                    mask[ndx] = entry[field] <= constraint[1]
                elif ineq == '==':
                    mask[ndx] = entry[field] == constraint[1]
                elif ineq == '>=':
                    mask[ndx] = entry[field] >= constraint[1]
                elif ineq == '>':
                    mask[ndx] = entry[field] > constraint[1]
                else:
                    raise Exception('Invalid inequality in constraint')
        else:
            for ndx, entry in enumerate(entries):
                mask[ndx] = entry[field] == constraint
        entries = self._mask(entries, mask)
        return entries
        
    def _deserialize(self, Cls, entry):
        """
        """
        obj = Cls()
        for k, v in entry.items():
            setattr(obj, k, v)
        return obj
        
    def create(self, entry):
        """Upon creation of an entry, new metadata is attached. An error is
           raised if metadata already exist.
        """
        if Meta.isAttached(entry):
            raise Exception("Metadata already attached to entry")
        table = self._getTableName(entry.__class__)
        meta = Meta()
        if not self._hasTable(table):
            ws = self._addTable(entry.__class__)
        else:
            ws = self.wb.get_sheet_by_name(table)
        meta._id, row_ndx = self._getNextId(ws)
        meta.attach(entry)
        self._write(ws, entry, row_ndx)
        return entry
        
    def read(self, Cls, filters={}):
        """
        """
        table = self._getTableName(Cls)
        ws = self.wb.get_sheet_by_name(table)
        entries = self._get(ws)
        for k, v in filters.items():
            entries = self._filter(entries, k, v)
        return [self._deserialize(Cls, entry) for entry in entries]
        
    def update(self, entry):
        """
        """
        pass
        
    def delete(self, entry):
        """
        """
        pass
